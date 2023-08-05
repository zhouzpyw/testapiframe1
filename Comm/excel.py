import openpyxl
'''
excel基本操作
'''
class excel:
    def __init__(self,filename,Sheetname):
        self.file = openpyxl.load_workbook(filename)
        self.Sheet = self.file[Sheetname]
        self.filename = filename

    def max_row(self):          #1.获取excel最大的行数
        max_row = self.Sheet.max_row
        return max_row

    def max_col(self):          #1.获取excel最大的列数
        max_col = self.Sheet.max_column
        return max_col

    def one_value(self,row_1,col_1):     #2.获取excel某个单元格的值
        value = self.Sheet.cell(row=row_1,column=col_1).value
        return  value

    def row_value(self,rownum):            #3.获取excel某行的值，以列表的方式输出
        data = []
        max_col = self.max_col()
        for i in range(0,max_col):
            values = self.Sheet.cell(row = rownum,column = i+1).value
            try:
                values=eval(values)
            except:
                pass
            finally:
                data.append(values)
        return data

    def value_row(self,rownum):          #4.获取excel某行的值，以字典的方式输出
        data = {}
        max_col = self.max_col()
        for i in range(1, max_col+1):
            key = self.Sheet.cell(row=1, column=i).value
            value = self.Sheet.cell(row=rownum, column=i).value
            try:
                value = eval(value)
            except:
                pass
            data[key] = value
        return data

    def value_col(self,colnum):          #4.获取excel某列的值，以字典的方式输出
        date={}
        max_row = self.Sheet.max_row
        for i in range(1,max_row+1):
            key = self.Sheet.cell(row=i,column = 1)
            value = self.Sheet.cell(row= i,column = colnum)
            date[key] = value
        return date

    def all_valuedict(self):          #5.获取excel的全部数据，列表字典输出[{},{},{}]
        date = []
        max_row = self.max_row()
        for i in range(2,max_row+1):
            rows = self.value_row(i)
            date.append(rows)
        return date

    def all_valuelist(self):  # 5.获取excel的全部数据，列表+列表输出[[],[],[]]
        date = []
        max_row = self.max_row()
        for i in range(2, max_row+1):
            rows = self.row_value(i)
            date.append(rows)
        return  date

    def write_value(self,rownum,colnum,values):     #6.在某个单元格中写入一个数据
        self.Sheet.cell(row=rownum,column=colnum).value=values
        try:
            self.file.save(self.filename)
        except Exception as e:
            print('保存失败',e)
        self.file.close()


if __name__ == '__main__':
    a=excel('F:\zzp\code_zzp\\testdata\info.xlsx','Sheet1')
    print(a.all_valuedict())



